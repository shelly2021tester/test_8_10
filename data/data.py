import openpyxl

class TestData:
    def __init__(self,file,sheetname):
        self.file=file
        self.wb=openpyxl.load_workbook(self.file)
        self.table=self.wb[sheetname]
        self.mrow=self.table.max_row
        self.mcol=self.table.max_column

    def read_xls(self):
        list_row=[]
        for row in range(2,self.mrow+1):
            list_col=[]
            for col in range(1,self.mcol+1):
                list_col.append(self.table.cell(row,col).value)
            list_row.append(list_col)
        return list_row

